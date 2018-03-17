'''
Created on Feb 20, 2018

@author: andersco
'''
from openpyxl import load_workbook
from time import sleep
from os import chdir


class parseAuditWorkbookClass(object):
    #public functions 
    def setAuditFilename(self, datalogFileName):
        self.__auditFileName = datalogFileName
        
        
    def getPatternComments(self):
        if not self.__dataPulled:
            self.__pullData()
        return self.__patternComments
    
    
    def getAnalogComments(self):
        if not self.__dataPulled:
            self.__pullData()
        return self.__analogComments
    
    
    def doNothing(self):
            sleep(0)  


    #private functions 
    def __init__(self):
        self.__auditFileName = "auditFile.xlsx"
        self.__dataPulled = False
        self.__patternComments = []
        self.__analogComments = []

    
    def __pullData(self):  
        if not self.__dataPulled:
            self.__openWorkbook()
            self.__getPatComments()
            self.__getAnalogComments()
            self.__wb.close()
            self.__dataPulled = True
  
   
    def __openWorkbook(self):
        self.__wb = load_workbook(self.__auditFileName)
        self.__wsPatterns=self.__wb["Patterns"]
        self.__wsAnalogTests=self.__wb["Analog Tests"]
       
    
    def __getPatComments(self):
        for row in self.__wsPatterns.rows:
            if (row[7]._value is not None) and (row[7]._value != "Comments"):
                vlctTestName = row[2].value
                vlctPatname = row[3].value
                uflexTestName = row[5].value
                uflexPatname = row[6].value
                comment = row[7].value
                commentTuple = (vlctTestName,vlctPatname,uflexTestName,uflexPatname,comment)
                self.__patternComments.append(commentTuple)


    def __getAnalogComments(self):
        for row in self.__wsAnalogTests.rows:
            if (row[15]._value is not None) and (row[15]._value != "Comments"):
                vlctTestName = row[2].value
                vlctMeasname = row[3].value
                uflexTestName = row[9].value
                uflexMeasname = row[10].value
                comment = row[15].value
                commentTuple = (vlctTestName,vlctMeasname,uflexTestName,uflexMeasname,comment)
                self.__analogComments.append(commentTuple)


#used to test class using runs as parseAuditWorkbookClass
if __name__ == '__main__':
    
    auditFile = parseAuditWorkbookClass()    
    auditFile.setAuditFilename("20180220_datalogCompare.xlsx")    
    chdir("..")
    patComments = auditFile.getPatternComments()
    auditFile.doNothing()
    
