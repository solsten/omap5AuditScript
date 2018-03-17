'''
Created on Jan 17, 2018

@author: andersco
'''
from datetime import date
from openpyxl import Workbook 
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from subModules.parseUflexDatalogClass import parseUflexDatalogClass
from subModules.parseVlctDatalogClass import parseVlctDatalogClass
from subModules.parseAuditWorkbookClass import parseAuditWorkbookClass

LINE_NUM = 0
FLOW_NUM = 0
VLCT_FLOW = 1
UFLEX_FLOW = 2
TEST_NAME = 1
PATTERN_NAME = 2
MEAS_NAME = 3
MIN_LIMIT = 4
MEAS_RSLT = 5
MAX_LIMIT = 6
UNITS = 7
PATTERN_SET = 8
NOT_FOUND = 99
VLCT_TEST = 2
VLCT_PAT = 3
UFLEX_TEST = 5
UFLEX_PAT = 6
DIG_COMMENT = 7

class compareVlctUflexDatalogsClass(object):

    #public functions
    def parseVlctDatalog(self, vlctDataLog):
        self.__vlct = parseVlctDatalogClass()
        self.__vlct.setDatalogFilename(vlctDataLog)
        self.__vlctTestNames = self.__vlct.getTestNames()
        self.__vlctAnalogMeasNames = self.__vlct.getAnalogMeasNames()
        self.__vlctPatternNames = self.__vlct.getPatternNames()
        self.__vlctContinuityTestNames = self.__vlct.getContinuityTests
        self.__vlctTestInstances = []
        self.__vlctTestInstancesUnsorted = self.__vlct.getTestInstances()
        self.__vlctTestInstancesUnsorted.sort(key=lambda tup: tup[LINE_NUM])
        for (i,testInstance) in enumerate(self.__vlctTestInstancesUnsorted):
            storeInst = tuple([i] + list(testInstance[1:]))
            self.__vlctTestInstances.append(storeInst)
        self.__wsVlctDatalog = self.__wb.create_sheet(title = "VLCT Datalog" )
        self.__wsVlctDatalog.append(self.__vlctDatalogHeader)
        self.__createVlctDatalog()
        self.__createExcelTable(self.__wsVlctDatalog,"vlctTable")
        del self.__vlct
 
    
    def parseUflexDatalog(self, vlctDataLog):
        self.__uflex = parseUflexDatalogClass()
        self.__uflex.setDatalogFilename(vlctDataLog)
        self.__uflexTestNames = self.__uflex.getTestNames()
        self.__uflexAnalogMeasNames = self.__uflex.getAnalogMeasNames()
        self.__uflexPatternNames = self.__uflex.getPatternNames()
        self.__uflexTestInstances = []
        self.__uflexTestInstancesUnsorted = self.__uflex.getTestInstances()
        self.__uflexTestInstancesUnsorted.sort(key=lambda tup: tup[0])
        for (i,testInstance) in enumerate(self.__uflexTestInstancesUnsorted):
            storeInst = tuple([i] + list(testInstance[1:]))            
            self.__uflexTestInstances.append(storeInst)
        self.__wsUflexDatalog = self.__wb.create_sheet(title = "UFLEX Datalog")
        self.__wsUflexDatalog.append(self.__uflexDatalogHeader)
        self.__createUflexDatalog()
        self.__createExcelTable(self.__wsUflexDatalog,"uflexTable")
        del self.__uflex
 
 
    def parseAuditFile(self,auditFileName):
            auditFile = parseAuditWorkbookClass()    
            auditFile.setAuditFilename(auditFileName)
            self.__patComments = self.__patComments + auditFile.getPatternComments()
            self.__analogComments = self.__analogComments + auditFile.getAnalogComments()
            auditFile.doNothing()

        
    def saveWorkbook(self):   
        today = date.today()
        workbookName = today.strftime("%Y%m%d") + "_datalogCompare.xlsx" 
        self.__wb.save(workbookName)
       

    def comparePatterns(self):
        self.__wsPatterns = self.__wb.create_sheet(title = "Patterns")
        self.__wsPatterns.append(self.__digitalHeader)        
        for vlctPatName in self.__vlctPatternNames:
            self.__getVlctDigitalTests(vlctPatName)
        for uflexPatName in self.__uflexPatternNames:
            self.__getUflexDigitalTests(uflexPatName)
        self.__attachDigitalComments()
        self.__createDigitalTestsSheet()
        self.__createExcelTable(self.__wsPatterns,"patternsTable")

   
    def compareAnalogTests(self):
        self.__wsAnalog = self.__wb.create_sheet(title = "Analog Tests")
        self.__wsAnalog.append(self.__analogHeader)        
        for measName in self.__vlctAnalogMeasNames:
            self.__getVlctAnalogTests(measName)
        for measName in self.__uflexAnalogMeasNames:
            self.__getUflexAnalogTests(measName)
        self.__attachAnalogComments()
        self.__createAnalogTestsSheet()
        self.__createExcelTable(self.__wsAnalog,"analogTable")

   
    def compareContinuityTests(self):
        self.__wsContinuity = self.__wb.create_sheet(title = "Opens Shorts Tests")
        self.__wsContinuity.append(self.__analogHeader)        
        for measName in self.__vlctContinuityTestNames:
            self.__getVlctContinuityTests(measName)
        for measName in self.__vlctContinuityTestNames:
            self.__getVlctContinuityTests(measName)
        self.__attachAnalogComments()
        self.__createAnalogTestsSheet()
        self.__createExcelTable(self.__wsAnalog,"analogTable")

   
    def __init__(self):      
        #openpyxl variables  
        self.__wb = Workbook()
        self.__wb.remove_sheet(self.__wb.active)
        
        #misc variables
        self.__patComments = []
        self.__analogComments = []
        self.__testFound = set()
        self.__knownPairs = []
        self.__digitalInstances = []
        self.__analogInstances = []
        self.__continuityInstances = []
        self.__patternSets = {}
        self.__parsedTestInstFile = False
        
        self.__analogHeader = ["Limits Match","VLCT Flow","VLCT Test Name","VLCT Meas Name",
                               "VLCT Min","VLCT Max","VLCT Meas Rslt","VLCT Unit","UFLEX Flow",
                                "UFLEX Test Name","UFLEX Meas Name","UFLEX Min","UFLEX Max",
                                "UFLEX Meas Rstl","UFLEX Unit","Comments"]
        
        self.__digitalHeader = ["Patterns Match","VLCT Flow","VLCT Test Name","VLCT Pattern",
                                    "UFLEX Flow","UFLEX Test Name","UFLEX Pattern","Comments"]
        
        self.__uflexDatalogHeader = ["Test Flow","Test Name","Pattern","Meas Name",
                                     "Min Limit", "Meas Rslt", "Max Limit","Units"]
        
        self.__vlctDatalogHeader = ["Test Flow","Test Name","Pattern","Meas Name",
                                     "Min Limit", "Meas Rslt", "Max Limit","Units"]
        
        self.__csiInitPats = ["SP_BG_ON","SD_ABEPLL_X4","SP_LDO_WKUP_EMU","BOOT_ANY","SP_SLDO_CORE_ACT",
                                "SP_SLDO_MPU_ACT","SD_COREPLL_X4","SP_SLDO_MM_ACT","SP_BBLDO_MPU_BYPASS",
                                "ST_ANA_CSI2","SP_BBLDO_MM_BYPASS","SP_SLDO_CORE_ACT_DISC","ST_ANA_CSI1",
                                "ST_ANA_CSI1","SP_SLDO_MM_ACT_DISC","SP_SLDO_MPU_ACT_DISC","XBCRXAB_LPRX_2DL",
                                "ST_ANA_CSI3","ST_ANA_DSIA","CSIA_V_HSRXVTH0VCM70_MO","SD_ABEPLL_X8",
                                "SD_COREPLL_X8","SD_SATAPLL_X125","SD_USB3PLL_X75","SP_SLDO_CORE_HIZ",
                                "SP_SLDO_MM_HIZ","SP_SLDO_MPU_HIZ","ST_ANA_DSIC","ST_ANA_SATA_RX",
                                "ST_ANA_USB3_RX","STDSIAC_BG_STRIM","STGRXAU_DLL_1P5_SEL","STGRXAS_DLL_2P5_SEL",
                                "STDSIAC_BG_STRIM","ST_ANA_HDMI","ST_TOP_AS","SP_BBLDO_MPU_OPNO","SP_ALL_ISORET_ON",
                                "SP_BBLDO_MM_OPNO","SP_ALL_CPU_INIT","BOOT_ATPG","STDSIAC_BG_STRIM"]
        
        self.__csiSkipInitTests = ["TP_CSIA_LPRXH_MO_ST","TP_CSIB_LPRXH_MO_ST","TP_CSIC_LPRXH_MO_ST",
                                   "TP_CSIA_ULPRXH_MO_ST","TP_CSIB_ULPRXH_MO_ST","TP_CSIC_ULPRXH_MO_ST",
                                   "TP_DSIA_LPRXH_MO_ST","TP_DSIA_ULPRXH_MO_ST","TP_DSIC_ULPRXH_MO_ST",
                                   "TP_DSIC_LPRXH_MO_ST","DSIA_V_LPRXVIH1Y_MO","DSIA_V_LPRXVIH3Y_MO",
                                   "CSIA_V_HSRXVTH0VCM200_MO","CSIA_V_HSRXVTH0VCM70_MO","CSIA_V_HSRXVTH4VCM200_MO",
                                   "CSIA_V_HSRXVTH4VCM70_MO","CSIB_V_CCPVTHAVCM06_MO","CSIB_V_CCPVTHAVCM09_MO",
                                   "CSIB_V_CCPVTHAVCM12_MO","CSIB_V_CCPVTHBVCM06_MO","CSIB_V_CCPVTHBVCM09_MO",
                                   "CSIB_V_HSRXVTH0VCM200_MO","CSIB_V_HSRXVTH0VCM70_MO","CSIB_V_HSRXVTH1VCM200_MO",
                                   "CSIB_V_HSRXVTH1VCM70_MO","CSIC_V_CCPVTHAVCM06_MO","CSIC_V_CCPVTHAVCM09_MO",
                                   "CSIC_V_CCPVTHAVCM12_MO","CSIC_V_CCPVTHBVCM06_MO","CSIC_V_CCPVTHBVCM09_MO",
                                   "CSIC_V_HSRXVTH0VCM200_MO","CSIC_V_HSRXVTH0VCM70_MO","CSIC_V_HSRXVTH1VCM200_MO",
                                   "CSIC_V_HSRXVTH1VCM70_MO","CSIA_V_LPRXVIH1Y_MO","CSIA_V_LPRXVIH3Y_MO",
                                   "CSIB_V_LPRXVIH1Y_MO","CSIC_V_LPRXVIH0Y_MO","DSIC_V_LPRXVIH2Y_MO",
                                   "DSIC_V_LPRXVIH3Y_MO","CSIA_V_ULPRXVIH1Y_MO","CSIA_V_ULPRXVIH3Y_MO",
                                   "CSIB_V_ULPRXVIH1Y_MO","CSIC_V_ULPRXVIH0Y_MO","DSIC_V_LPRXVIH1Y_MO",
                                   "DSIC_V_ULPRXVIH2Y_MO","DSIC_V_ULPRXVIH3Y_MO","TP_DSIC_LPRXH_MO_ST"
                                   "DSIC_V_LPRXVIH1Y_MO","DSIC_V_LPRXVIH2Y_MO","DSIC_V_LPRXVIH3Y_MO",
                                   "TP_DSIC_ULPRXH_MO_ST","DSIC_V_ULPRXVIH1Y_MO","DSIC_V_ULPRXVIH1Y_MO",
                                   "DSIC_V_ULPRXVIH3Y_MO","USB2_HSRX_THVCM2_ST","USB2_HSRX_THVCM3_ST",
                                   "DSIA_V_ULPRXVIH1Y_MO","DSIA_V_ULPRXVIH3Y_MO","TP_DSIA_LPCDH_MO_ST",
                                   "DSIA_V_LPCDVIH1Y_MO","DSIA_V_LPCDVIH3Y_MO","TP_DSIC_LPCDH_MO_ST","DSIC_V_LPCDVIH1Y_MO",
                                   "DSIC_V_LPCDVIH2Y_MO","DSIC_V_LPCDVIH3Y_MO","USB2_HSRX_SQVCM2_ST","USB2_HSRX_SQVCM3_ST",
                                   "USB2_CEPUDP_ST","USB2_GPIOTXCURR_ST","USB2_PADLEAKDM_ST","U3RX_V_RXTHRESHOLD1_NO",
                                   "U3RX_V_RXTHRESHOLD2_NO","U3RX_V_RXTHRESHOLD3_NO","SARX_V_RXTHRESHOLD1_NO","SARX_V_RXTHRESHOLD2_NO",
                                   "SARX_V_RXTHRESHOLD3_NO","TP_DSIC_LPRXL_MO_ST","TP_CSIC_ULPRXL_MO_ST","HDMI_RX_DETECT_MO_ST",
                                   "SARX_TRIM_LOSD_ST","U3RX_TRIM_LOSD_ST","MP_RETIVA1_TC_OPNO_ST","MP_RETIVA2_TC_OPNO_ST",
                                   "MB_STM_BYG_TC_OPLO_MEM_ST","HDMI_TXDC_ST","IQ_VIDEOPB_NOZ_WKUP_ST","IQ_VIDEOPB_NOZ_WKUP_ST",
                                   "TP_DSIC_LPCDL_MO_ST","TP_DSIC_ULPRXL_MO_ST"]
        
        self.__1stInstOnlyPats = ["CSIA_V_HSRXVTH0VCM70_MO","CSIA_V_HSRXVTH0VCM200_MO","CSIA_V_HSRXVTH0VCM330_MO",
                                 "CSIA_V_HSRXVTH1VCM70_MO","CSIA_V_HSRXVTH1VCM200_MO","CSIA_V_HSRXVTH1VCM330_MO",
                                 "CSIA_V_HSRXVTH2VCM70_MO","CSIA_V_HSRXVTH2VCM200_MO","CSIA_V_HSRXVTH2VCM330_MO",
                                 "CSIA_V_HSRXVTH3VCM70_MO","CSIA_V_HSRXVTH3VCM200_MO","CSIA_V_HSRXVTH3VCM330_MO",
                                 "CSIA_V_HSRXVTH4VCM70_MO","CSIA_V_HSRXVTH4VCM200_MO","CSIA_V_HSRXVTH4VCM330_MO",
                                 "CSIB_V_HSRXVTH0VCM70_MO","CSIB_V_HSRXVTH0VCM200_MO","CSIB_V_HSRXVTH0VCM330_MO",
                                 "CSIB_V_HSRXVTH1VCM70_MO","CSIB_V_HSRXVTH1VCM200_MO","CSIB_V_HSRXVTH1VCM330_MO",
                                 "CSIB_V_HSRXVTH2VCM70_MO","CSIB_V_HSRXVTH2VCM200_MO","CSIB_V_HSRXVTH2VCM330_MO",
                                 "CSIC_V_HSRXVTH0VCM70_MO","CSIC_V_HSRXVTH0VCM200_MO","CSIC_V_HSRXVTH0VCM330_MO",
                                 "CSIC_V_HSRXVTH1VCM70_MO","CSIC_V_HSRXVTH1VCM200_MO","CSIC_V_HSRXVTH1VCM330_MO",
                                 "CSIA_V_CCPVTHAVCM06_MO","CSIA_V_CCPVTHBVCM06_MO","CSIA_V_CCPVTHAVCM09_MO","CSIA_V_CCPVTHBVCM09_MO",
                                 "CSIA_V_CCPVTHAVCM12_MO","CSIA_V_CCPVTHBVCM12_MO","CSIB_V_CCPVTHAVCM06_MO","CSIB_V_CCPVTHBVCM06_MO",
                                 "CSIB_V_CCPVTHAVCM09_MO","CSIB_V_CCPVTHBVCM09_MO","CSIB_V_CCPVTHAVCM12_MO","CSIB_V_CCPVTHBVCM12_MO",
                                 "CSIC_V_CCPVTHAVCM06_MO","CSIC_V_CCPVTHBVCM06_MO","CSIC_V_CCPVTHAVCM09_MO","CSIC_V_CCPVTHBVCM09_MO",
                                 "CSIC_V_CCPVTHAVCM12_MO","CSIC_V_CCPVTHBVCM12_MO","CSIA_V_LPRXVIL0X_MO","CSIA_V_LPRXVIL0Y_MO",
                                 "CSIA_V_LPRXVIL1X_MO","CSIA_V_LPRXVIL1Y_MO","CSIA_V_LPRXVIL2X_MO","CSIA_V_LPRXVIL2Y_MO",
                                 "CSIA_V_LPRXVIL3X_MO","CSIA_V_LPRXVIL3Y_MO","CSIA_V_LPRXVIL4X_MO","CSIA_V_LPRXVIL4Y_MO",
                                 "CSIB_V_LPRXVIL0X_MO","CSIB_V_LPRXVIL0Y_MO","CSIB_V_LPRXVIL1X_MO","CSIB_V_LPRXVIL1Y_MO",
                                 "CSIB_V_LPRXVIL2X_MO","CSIB_V_LPRXVIL2Y_MO","CSIC_V_LPRXVIL0X_MO","CSIC_V_LPRXVIL0Y_MO",
                                 "CSIC_V_LPRXVIL1X_MO","CSIC_V_LPRXVIL1Y_MO","CSIA_V_ULPRXVIL0X_MO","CSIA_V_ULPRXVIL0Y_MO",
                                 "CSIA_V_ULPRXVIL1X_MO","CSIA_V_ULPRXVIL1Y_MO","CSIA_V_ULPRXVIL2X_MO","CSIA_V_ULPRXVIL2Y_MO",
                                 "CSIA_V_ULPRXVIL3X_MO","CSIA_V_ULPRXVIL3Y_MO","CSIA_V_ULPRXVIL4X_MO","CSIA_V_ULPRXVIL4Y_MO",
                                 "CSIB_V_ULPRXVIL0X_MO","CSIB_V_ULPRXVIL0Y_MO","CSIB_V_ULPRXVIL1X_MO","CSIB_V_ULPRXVIL1Y_MO",
                                 "CSIB_V_ULPRXVIL2X_MO","CSIB_V_ULPRXVIL2Y_MO","CSIC_V_ULPRXVIL0X_MO","CSIC_V_ULPRXVIL0Y_MO",
                                 "CSIC_V_ULPRXVIL1X_MO","CSIC_V_ULPRXVIL1Y_MO","DSIA_V_LPRXVIL0X_MO","DSIA_V_LPRXVIL0Y_MO",
                                 "DSIA_V_LPRXVIL1X_MO","DSIA_V_LPRXVIL1Y_MO","DSIA_V_LPRXVIL2X_MO","DSIA_V_LPRXVIL2Y_MO",
                                 "DSIA_V_LPRXVIL3X_MO","DSIA_V_LPRXVIL3Y_MO","DSIA_V_LPRXVIL4X_MO","DSIA_V_LPRXVIL4Y_MO",
                                 "DSIC_V_LPRXVIL0X_MO","DSIC_V_LPRXVIL0Y_MO","DSIC_V_LPRXVIL1X_MO","DSIC_V_LPRXVIL1Y_MO",
                                 "DSIC_V_LPRXVIL2X_MO","DSIC_V_LPRXVIL2Y_MO","DSIC_V_LPRXVIL3X_MO","DSIC_V_LPRXVIL3Y_MO",
                                 "DSIC_V_LPRXVIL4X_MO","DSIC_V_LPRXVIL4Y_MO","DSIA_V_ULPRXVIL0X_MO","DSIA_V_ULPRXVIL0Y_MO",
                                 "DSIA_V_ULPRXVIL1X_MO","DSIA_V_ULPRXVIL1Y_MO","DSIA_V_ULPRXVIL2X_MO","DSIA_V_ULPRXVIL2Y_MO",
                                 "DSIA_V_ULPRXVIL3X_MO","DSIA_V_ULPRXVIL3Y_MO","DSIA_V_ULPRXVIL4X_MO","DSIA_V_ULPRXVIL4Y_MO",
                                 "DSIC_V_ULPRXVIL0X_MO","DSIC_V_ULPRXVIL0Y_MO","DSIC_V_ULPRXVIL1X_MO","DSIC_V_ULPRXVIL1Y_MO",
                                 "DSIC_V_ULPRXVIL2X_MO","DSIC_V_ULPRXVIL2Y_MO","DSIC_V_ULPRXVIL3X_MO","DSIC_V_ULPRXVIL3Y_MO",
                                 "DSIC_V_ULPRXVIL4X_MO","DSIC_V_ULPRXVIL4Y_MO","DSIA_V_LPCDVIL0X_MO","DSIA_V_LPCDVIL0Y_MO",
                                 "DSIA_V_LPCDVIL1X_MO","DSIA_V_LPCDVIL1Y_MO","DSIA_V_LPCDVIL2X_MO","DSIA_V_LPCDVIL2Y_MO",
                                 "DSIA_V_LPCDVIL3X_MO","DSIA_V_LPCDVIL3Y_MO","DSIA_V_LPCDVIL4X_MO","DSIA_V_LPCDVIL4Y_MO",
                                 "DSIC_V_LPCDVIL0X_MO","DSIC_V_LPCDVIL0Y_MO","DSIC_V_LPCDVIL1X_MO","DSIC_V_LPCDVIL1Y_MO",
                                 "DSIC_V_LPCDVIL2X_MO","DSIC_V_LPCDVIL2Y_MO","DSIC_V_LPCDVIL3X_MO","DSIC_V_LPCDVIL3Y_MO",
                                 "DSIC_V_LPCDVIL4X_MO","DSIC_V_LPCDVIL4Y_MO","USB2_V_FSRXTHDP_MO","USB2_V_FSRXTHDM_MO",
                                 "USB2_V_FSRXDIFFTHVCM1_MO","USB2_V_FSRXDIFFTHVCM2_MO","USB2_V_FSRXDIFFTHVCM3_MO","USB2_V_HSTXDISCTHDP_MO",
                                 "USB2_V_HSTXDISCTHDM_MO","USB2_V_HSRX_THVCM1_MO","USB2_V_HSRX_THVCM2_MO","USB2_V_HSRX_THVCM3_MO",
                                 "USB2_V_HSRX_SQVCM1_MO","USB2_V_HSRX_SQVCM2_MO","USB2_V_HSRX_SQVCM3_MO","USB2_V_CECOMPTH_MO",
                                 "USB2_V_CECOMP2THDP_MO","USB2_V_CECOMP2THDM_MO","USB2_V_GPIORXDP_MO","USB2_V_GPIORXDM_MO",
                                 "SARX_V_LOSDPOSTTRIM_NO","SARX_V_LOSDPOSTTRIM2_NO","SARX_V_RXTHRESHOLD1_NO","SARX_V_RXTHRESHOLD2_NO",
                                 "SARX_V_RXTHRESHOLD3_NO","SARX_V_RXTHRESHOLD4_NO","U3RX_V_RXTHRESHOLD1_MO","U3RX_V_RXTHRESHOLD3_MO",
                                 "U3RX_V_LOSDPOSTTRIM2_MO","U3RX_V_RXTHRESHOLD2_MO","U3RX_V_RXTHRESHOLD4_MO","U3RX_V_LOSDPOSTTRIM_MO"]
              
        self.__patStartedInVba = ["XBCRXAA_BG_TRIM","XBCRXAB_BG_TRIM","XBCRXAC_BG_TRIM","XBCRXAA_RTERM_TRIM","XBCRXAB_RTERM_TRIM",
                                  "XBCRXAC_RTERM_TRIM","XBDSIAA_BG_TRIM","XBDSIAC_BG_TRIM","XBDSIAA_RTERM_TRIMD","XBDSIAC_RTERM_TRIMD",
                                  "XSDSIAA_DLY_TRIM","XSDSIAC_DLY_TRIM","XBHDMA_BG_TRIM","XBHDMA_SC_TRIM","XBGRXAU_BG_TRIM",
                                  "XBGRXAU_RTERM_TRIM","XSGRXAU_DLLOFF_TRIM","XSGRXAU_DLL0P7_TRIM","XSGRXAU_DLL1P5_TRIM","XSGRXAU_DLL2P5_TRIM",
                                  "XBGRXAU_LOSD_TRIM","XBGRXAU_LOSD_TRIM","XBGRXAS_BG_TRIM","STGRXAS_BG_STRIM","XBGRXAS_RTERM_TRIM",
                                  "XSGRXAS_DLLOFF_TRIM","XSGRXAS_DLL0P7_TRIM","XSGRXAS_DLL1P5_TRIM","XBGRXAS_LOSD_TRIM","XBGRXAS_LOSD_TRIM",
                                  "XSSTXAU_BG_TRIM","XSSTXAU_RTERM_TRIM","XSSTXAS_BG_TRIM","XSSTXAS_RTERM_TRIM","XSUS2A_BG_TRIM1",
                                  "XSUS2A_BG_TRIM2","XSUS2A_RTERM_TRIM","XBHDMA_TXDC_CHAR0","XBHDMA_TXDC_CHAR0","XBHDMA_TXDC_CHAR0",
                                  "XBHDMA_TXDC_CHAR0","XSGRXASSAMPOFMEASPM1","XSGRXASSAMPOFMEASPM2","XSGRXASSAMPOFMEASPM3",
                                  "XSGRXASSAMPOFMEASPM4","XBCRXAAHSRX0_4DL","XBCRXAAHSRX1_4DL","XBCRXAAHSRX2_4DL",
                                  "XBCRXAAHSRX3_4DL","XBCRXAAHSRX1_4DL","XBCRXAALPRX_4DL","XBCRXAALPRX_4DL","XBDSIAALPCD_4DL","XBCRXAAHSRX4_4DL",
                                  "XBCRXACULPRX1DL","XBDSIAAULPRX4DL","XBCRXAAULPRX4DL","XBDSIAALPRX_4DL","XBCRXACLPRX_1DL","XBDSIAA_RTERM_TRIM",
                                  "XBDSIAC_BIAS_MEAS","XBGRXASLOSD_MEAS","XBCRXAALPRX_4DL","XBCRXACLPRX_1DL","XBDSIAA_BIAS_MEAS","XBDSIAALPCD_4DL",
                                  "XBDSIAALPRX_4DL","XBDSIAA_RTERM_TRIM","XBGRXAULOSD_MEAS","XBDSIAC_RTERM_TRIM","XSUS2A_RTERM_MEAS",
                                  "XSGRXAS_DLL2P5_TRIM","STGRXAS_RTERM_STRIM","XSGRXAUSAMPOFMEASPM1","XSGRXAUSAMPOFMEASPM2","XSGRXAUSAMPOFMEASPM3",
                                  "XSGRXAUSAMPOFMEASPM4","STGRXAU_BG_STRIM","STGRXAU_RTERM_STRIM","ST_ANA_USB3_RX","ST_ANA_SATA_RX"
                                  ]
                
        self.__pairedTests = {"BOOT_ATPG_RMR_TC_VMAX_ST":"BOOT_ATPG_TC_VMAX_ST",
              "BLDO2_PRETRM_":"BUILD_VBBTABLE_FUNC",
              "EFUSE_STD_RUNAUTOLOAD_1_ST":"EFUSECTLRRUNAUTOLOADCHECK",
              "IQ_IO_HSIC_ALL0_ST":"IDDQ_IOMODE_2",
              "IQ_PD_CEFUSE_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_STDBY4_JUZ_WKUPL_ST":"IDDQ_IOMODE_1",
              "MP_RETDSS1_TC_OPNO_ST":"MP_DSS1_RET_TC_OPNO_CORE_ST",
              "MP_RETDSS2_TC_OPNO_ST":"MP_DSS2_RET_TC_OPNO_CORE_ST",
              "MP_RETGPHY1_TC_OPNO_ST":"MP_GPHY1_RET_TC_OPNO_ST",
              "MP_RETGPHY2_TC_OPNO_ST":"MP_GPHY2_RET_TC_OPNO_ST",
              "MP_RETIVA1_TC_OPNO_ST":"MP_IVA1_RET_TC_OPNO_ST",
              "MP_RETIVA2_TC_OPNO_ST":"MP_IVA2_RET_TC_OPNO_ST",
              "MP_RETTSLA1_TC_OPNO_ST":"MP_TSLA1_RET_TC_OPNO_ST",
              "MP_RETTSLA2_TC_OPNO_ST":"MP_TSLA2_RET_TC_OPNO_ST",
              "MP_RETZONDA1_TC_OPNO_ST":"MP_ZONDA1_RET_TC_OPNO_ST",
              "MP_RETZONDA2_TC_OPNO_ST":"MP_ZONDA2_RET_TC_OPNO_ST",
              "MP_RETZONDA3_TC_OPNO_ST":"MP_ZONDA3_RET_TC_OPNO_ST",
              "MP_RETZONDA4_TC_OPNO_ST":"MP_ZONDA4_RET_TC_OPNO_ST",
              "PTRIM_TSMM_ST":"PTRIM_TS_VNOM",
              "PTRIM_TSMPU_ST":"PTRIM_TS_VNOM",
              "PTRIM_VLDO1":"PTRIMVSLDO1_VNOM",
              "PTRIM_VLDO2":"PTRIMVSLDO2_VNOM",
              "PTRIM_VLDO3":"PTRIMVSLDO3_VNOM",
              "PTRIM_VLDO4":"PTRIMVSLDO4_VNOM",
              "TG_BBLDOMM_VSSA_ST":"TRIM_BBMM_VSS_ST",
              "TG_BBLDOMPU_VSSA_ST":"TRIM_BBMPU_VSS_ST",
              "TG_PTRIM1_VBGAP1_ST":"PTRIM_VBGAPTSMPU_ST",
              "TG_PTRIM1_VBGAP2_ST":"PTRIM_VBGAPTSMM_ST",
              "TG_PTRIM1_VBGAP3_MI_ST":"PTRIM_VBGAPTSCORE_MI_ST",
              "TG_PTRIM1_VBGAP3_ST":"PTRIM_VBGAPTSCORE_ST",
              "TG_PTRIM1_VBGAP4_ST":"PTRIM_VBGAPTSMPU_2_ST",
              "TG_PTRIM1_VBGAP5_ST":"PTRIM_VBGAPTSMPU_3_ST",
              "TG_PTRIM2_VBGAP1_ST":"PTRIM_VBGAPTSMPU2_ST",
              "TG_PTRIM2_VBGAP2_ST":"PTRIM_VBGAPTSMM2_ST",
              "TG_PTRIM2_VBGAP3_MI_ST":"PTRIM_VBGAPTSCOREI_MI_ST",
              "TG_PTRIM2_VBGAP3_ST":"PTRIM_VBGAPTSCOREI_ST",
              "TG_PTRIM2_VBGAP4_ST":"PTRIM_VBGAPTSMPU_22_ST",
              "TG_PTRIM2_VBGAP5_ST":"PTRIM_VBGAPTSMPU_32_ST",
              "TG_PTRIMVSSA_VBGAP1_ST":"PTRIM_BGTS1_VSSA_ST",
              "TG_PTRIMVSSA_VBGAP2_ST":"PTRIM_BGTS2_VSSA_ST",
              "TG_PTRIMVSSA_VBGAP3_MI_ST":"PTRIM_BGTS3_VSSA_MI_ST",
              "TG_PTRIMVSSA_VBGAP3_ST":"PTRIM_BGTS3_VSSA_ST",
              "TG_PTRIMVSSA_VBGAP4_ST":"PTRIM_BGTS4_VSSA_ST",
              "TG_PTRIMVSSA_VBGAP5_ST":"PTRIM_BGTS5_VSSA_ST",
              "TG_RNG_CLKOUT_ST":"RNG_CLKOUT_ST",
              "TG_RNG_GNG_ST":"RNG_B_GNG_MO_ST",
              "TG_U3TX_RXDET1_CM2_ST":"U3TX_RXDET1_CM2_ST",
              "TG_WLDO_ACT_M1_ST":"WKUPLDO_ACT_M1_ST",
              "TG_WLDO_ACT_MX_ST":"WKUPLDO_ACT_MX_ST",
              "TG_WLDO_EMU_MI_ST":"WKUPLDO_EMU_MI_ST",
              "TG_WLDO_EMU_MX_ST":"WKUPLDO_EMU_MX_ST",
              "TG_WLDO_SLEEP_M1_ST":"WKUPLDO_SLEEP_M1_ST",
              "TG_WLDO_SLEEP_MX_ST":"WKUPLDO_SLEEP_MX_ST",
              "TG_USB2_GPIORX_ST":"USB2_GPIORX_ST",
              "TP_CSIC_HSRX_VTH1_MO_ST":"CSIC_V_HSRXVTH1VCM",
              "TP_DSIA_HSTX_0_MO_ST":"TG_DSIA_HSTX_EMUL_MO",
              "TP_DSIC_HSTX_0_MO_ST":"TG_DSIC_HSTX_EMUL_MO",
              "TRIM_BBMPU_VSS_ST":"BUILD_VBBTABLE_FUNC",
              "USB2_GPIORXCURR_ST":"USB2_GPIOTXCURR_ST",
              "XS_DPLA_ABE_TC_OPNO_CORE_104_ST":"XS_DPLA_ABE_104_TC_OPNO_CORE_ST",
              "XS_DPLA_ABE_TC_OPNO_CORE_10M_ST":"XS_DPLA_ABE_10M_TC_OPNO_CORE_ST",
              "XS_DPLA_ABE_TC_OPNO_CORE_20M_ST":"XS_DPLA_ABE_20M_TC_OPNO_CORE_ST",
              "XS_DPLA_ABE_TC_OPNO_CORE_300K_ST" :"XS_DPLA_ABE_300K_TC_OPNO_CORE_ST",
              "XS_DPLA_ABE_TC_OPNO_CORE_52_ST":"XS_DPLA_ABE_52_TC_OPNO_CORE_ST",
              "XS_DPLA_CORE_TC_OPNO_CORE_104_ST":"XS_DPLA_CORE_104_TC_OPNO_CORE_ST",
              "XS_DPLA_CORE_TC_OPNO_CORE_10M_ST":"XS_DPLA_CORE_10M_TC_OPNO_CORE_ST",
              "XS_DPLA_CORE_TC_OPNO_CORE_20M_ST":"XS_DPLA_CORE_20M_TC_OPNO_CORE_ST",
              "XS_DPLA_CORE_TC_OPNO_CORE_300K_ST":"XS_DPLA_CORE_300K_TC_OPNO_CORE_ST",
              "XS_DPLA_CORE_TC_OPNO_CORE_52_ST":"XS_DPLA_CORE_52_TC_OPNO_CORE_ST",
              "XS_DPLA_DBG_TC_OPNO_CORE_104_ST":"XS_DPLA_DBG_104_TC_OPNO_CORE_ST",
              "XS_DPLA_DBG_TC_OPNO_CORE_10M_ST":"XS_DPLA_DBG_10M_TC_OPNO_CORE_ST",
              "XS_DPLA_DBG_TC_OPNO_CORE_20M_ST":"XS_DPLA_DBG_20M_TC_OPNO_CORE_ST",
              "XS_DPLA_DBG_TC_OPNO_CORE_300K_ST":"XS_DPLA_DBG_300K_TC_OPNO_CORE_ST",
              "XS_DPLA_DBG_TC_OPNO_CORE_52_ST":"XS_DPLA_DBG_52_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIA_TC_OPNO_CORE_104_ST":"XS_DPLA_DSIA_104_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIA_TC_OPNO_CORE_10M_ST":"XS_DPLA_DSIA_10M_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIA_TC_OPNO_CORE_20M_ST":"XS_DPLA_DSIA_20M_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIA_TC_OPNO_CORE_300K_ST":"XS_DPLA_DSIA_300K_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIA_TC_OPNO_CORE_52_ST":"XS_DPLA_DSIA_52_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIC_TC_OPNO_CORE_104_ST":"XS_DPLA_DSIC_104_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIC_TC_OPNO_CORE_10M_ST":"XS_DPLA_DSIC_10M_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIC_TC_OPNO_CORE_20M_ST":"XS_DPLA_DSIC_20M_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIC_TC_OPNO_CORE_300K_ST":"XS_DPLA_DSIC_300K_TC_OPNO_CORE_ST",
              "XS_DPLA_DSIC_TC_OPNO_CORE_52_ST":"XS_DPLA_DSIC_52_TC_OPNO_CORE_ST",
              "XS_DPLA_DSP_TC_OPNO_CORE_104_ST":"XS_DPLA_DSP_104_TC_OPNO_CORE_ST",
              "XS_DPLA_DSP_TC_OPNO_CORE_10M_ST":"XS_DPLA_DSP_10M_TC_OPNO_CORE_ST",
              "XS_DPLA_DSP_TC_OPNO_CORE_20M_ST":"XS_DPLA_DSP_20M_TC_OPNO_CORE_ST",
              "XS_DPLA_DSP_TC_OPNO_CORE_300K_ST":"XS_DPLA_DSP_300K_TC_OPNO_CORE_ST",
              "XS_DPLA_DSP_TC_OPNO_CORE_52_ST":"XS_DPLA_DSP_52_TC_OPNO_CORE_ST",
              "XS_DPLA_MPU_TC_OPNO_CORE_104_ST":"XS_DPLA_MPU_104_TC_OPNO_CORE_ST",
              "XS_DPLA_MPU_TC_OPNO_CORE_10M_ST":"XS_DPLA_MPU_10M_TC_OPNO_CORE_ST",
              "XS_DPLA_MPU_TC_OPNO_CORE_20M_ST":"XS_DPLA_MPU_20M_TC_OPNO_CORE_ST",
              "XS_DPLA_MPU_TC_OPNO_CORE_300K_ST":"XS_DPLA_MPU_300K_TC_OPNO_CORE_ST",
              "XS_DPLA_MPU_TC_OPNO_CORE_52_ST":"XS_DPLA_MPU_52_TC_OPNO_CORE_ST",
              "XS_DPLA_PER_TC_OPNO_CORE_104_ST":"XS_DPLA_PER_104_TC_OPNO_CORE_ST",
              "XS_DPLA_PER_TC_OPNO_CORE_10M_ST":"XS_DPLA_PER_10M_TC_OPNO_CORE_ST",
              "XS_DPLA_PER_TC_OPNO_CORE_20M_ST":"XS_DPLA_PER_20M_TC_OPNO_CORE_ST",
              "XS_DPLA_PER_TC_OPNO_CORE_300K_ST":"XS_DPLA_PER_300K_TC_OPNO_CORE_ST",
              "XS_DPLA_PER_TC_OPNO_CORE_52_ST":"XS_DPLA_PER_52_TC_OPNO_CORE_ST",             
              "XS_ALLIPPING_TC_OPNO_ST":"XS_TOP_MPUALLIP_TC_OPNO_CORE_ST"}

        self.__pairedPatterns = {"ARTOP_BSU_0":"ARTOP_BSU_0_MOD",
            "BOOT_ATPG_SYSNRESWARM":"BOOT_ATPG",
            "BOOT_ATPG_SYSNRESWARM":"BOOT_ATPG",
            "OMAP5_XI_OSC_FREERUN":"OMAP5_XI_OSC_FR",
            "SATX_HALFSWINGGEN1_2_ST":"PTRIM_TS_VNOM",
            "SP_SLDO_CORE_ACT":"SP_SLDO_CORE_ACT_DISC",
            "SP_SLDO_MM_ACT":"SP_SLDO_MM_ACT_DISC",
            "SP_SLDO_MPU_ACT":"SP_SLDO_MPU_ACT_DISC",
            "XSGRXAS_1P5_LPBK_ALDO_9":"XSGRXAS_1P5_LPBK",
            "XSGRXAU_1P5_LPBK_ALDO_9":"XSGRXAS_1P5_LPBK",
            "XSTOP_DPHY1_CPE":"XSTOP_DPHY1_CPE_MOD2",
            "XSTOP_DPHY1_LFPE":"XSTOP_DPHY1_LFPE_MOD2",
            "XSTOP_DPHY2_CPE":"XSTOP_DPHY2_CPE_MOD2",
            "XSTOP_DPHY2_LFPE":"XSTOP_DPHY2_LFPE_MOD2",
            "XSUS2A_CAL_READ":"XSUS2A_CAL_READ_MODIFY",
            "XSUS2A_RTERM_MEAS":"XSUS2A_RTERM_MEAS_MODIFIED"}
        
        self.__iatopPairedTests = {"IQ_PD_ABE_MPU_ST":"IDDQ_PER_PD_1",
              "IQ_PD_ABE_L_ST":"IDDQ_PER_PD_1",
              "IQ_PD_CAM_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_CAM_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_CEFUSE_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_CORE_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_COREAON_MPU_ST":"IDDQ_PER_PD_2",
              "IQ_PD_DSP_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_DSS_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_EMU_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_GPU0H_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_GPU1H_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_GPUH_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_IVA_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_L3INIT_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_L3INIT_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU0N_MPU_HI_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU0N_MPU_LO_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPUN0_MPU_ST":"IDDQ_PER_PD_4",
              "IQ_PD_MPU0N_MPU_MID_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU0N_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU1N_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPUN_MPU_ST":"IDDQ_PER_PD_3",
              "IQ_PD_CORE_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_COREAON_L_ST":"IDDQ_PER_PD_2",
              "IQ_PD_DSP_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_DSS_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_EMU_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_GPU0H_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_GPU1H_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_GPUH_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_IVA_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_L3INIT_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_L3INIT_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU0N_L_HI_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU0N_L_LO_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU0N_L_MID_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU0N_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPU1N_L_ST":"IDDQ_PER_PD_3",
              "IQ_PD_MPUN_L_ST":"IDDQ_PER_PD_3",
              "IQ_STDBY1_JUZ_CORE_ST":"IDDQ_STANDBY_1",
              "IQ_STDBY4_JUZ_CORE_ST":"IDDQ_STANDBY_2",
              "IQ_VIDEOPB_NOZ_CORE_ST":"IDDQ_VIDEOPLAYBACK",
              "IQ_PD_TEST_L_ST":"IDDQ_PER_PD_5"}
        

    def __getVlctAnalogTests(self,measName):
        vlctInst = next((inst for inst in self.__vlctTestInstances if measName in inst), NOT_FOUND)
        if measName in self.__uflexAnalogMeasNames:
            uflexInst = next((inst for inst in self.__uflexTestInstances if measName in inst), NOT_FOUND)
            if uflexInst != NOT_FOUND:
                newRow = self.__getVlctUflexRow(vlctInst,uflexInst)
                self.__analogInstances.append(newRow)
                self.__uflexAnalogMeasNames.remove(measName)
        else:
            newRow = self.__getVlctRow(vlctInst)
            self.__analogInstances.append(newRow)

    
    def __getVlctContinuityTests(self,measName):
        vlctInst = next((inst for inst in self.__vlctTestInstances if measName in inst), NOT_FOUND)
        #if measName in self.__uflexAnalogMeasNames:
            #uflexInst = next((inst for inst in self.__uflexTestInstances if measName in inst), NOT_FOUND)
            #if uflexInst != NOT_FOUND:
                #newRow = self.__getVlctUflexRow(vlctInst,uflexInst)
                #self.__continuityInstances.append(newRow)
                #self.__uflexAnalogMeasNames.remove(measName)
        #else:
        newRow = self.__getVlctRow(vlctInst)
        self.__continuityInstances.append(newRow)

    
    def __getUflexAnalogTests(self,measName):
        uflexInst = next((inst for inst in self.__uflexTestInstances if measName in inst[MEAS_NAME]), NOT_FOUND)
        if uflexInst != NOT_FOUND:
            newRow = self.__getUflexRow(uflexInst)
            self.__analogInstances.append(newRow)

    
    def __getUflexRow(self,uflexInst):
        if "PINSHORTSTEST" in uflexInst[TEST_NAME]:
            newRow = ("YES","","","","","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexInst[MEAS_NAME],
                  uflexInst[MIN_LIMIT],uflexInst[MAX_LIMIT],uflexInst[MEAS_RSLT],uflexInst[UNITS],
                  "Pin Shorts datalogged by group on VLCT and by pin on UFlex")
        elif "_HSTX_EMUL_MO" in uflexInst[TEST_NAME]:
            newRow = ("YES","","","","","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexInst[MEAS_NAME],
                  uflexInst[MIN_LIMIT],uflexInst[MAX_LIMIT],uflexInst[MEAS_RSLT],uflexInst[UNITS],
                  "meas not datalogged in VLCT program")
        elif "EFUSEAUDIT" in uflexInst[TEST_NAME]:
            newRow = ("YES","","","","","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexInst[MEAS_NAME],
                  uflexInst[MIN_LIMIT],uflexInst[MAX_LIMIT],uflexInst[MEAS_RSLT],uflexInst[UNITS],
                  "This EFUSE not checked on VLCT")
        elif "PRETRM_" in uflexInst[MEAS_NAME]:
            newRow = ("YES","","","","","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexInst[MEAS_NAME],
                  uflexInst[MIN_LIMIT],uflexInst[MAX_LIMIT],uflexInst[MEAS_RSLT],uflexInst[UNITS],
                  "exact trim values change from device to device")
        elif "CVM_" in uflexInst[MEAS_NAME]:
            newRow = ("YES","","","","","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexInst[MEAS_NAME],
                  uflexInst[MIN_LIMIT],uflexInst[MAX_LIMIT],uflexInst[MEAS_RSLT],uflexInst[UNITS],
                  "setup test, no limits")
        else:
            newRow = ("NO","","","","","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexInst[MEAS_NAME],
                  uflexInst[MIN_LIMIT],uflexInst[MAX_LIMIT],uflexInst[MEAS_RSLT],uflexInst[UNITS],"")
        return newRow
    

    def __getVlctDigitalTests(self,vlctPatName):
        #get all the VLCT test instances that use this pattern   
        vlctTestInstances = [inst for inst in self.__vlctTestInstances if vlctPatName in inst]
        for vlctInst in vlctTestInstances:
            if "PTRIM_ADPLLABELOCK1_ST" in vlctInst[TEST_NAME]:
                self.__doNothing()
            uflexPatName = self.__getUflexPatname(vlctPatName)
            if uflexPatName in self.__uflexPatternNames:
                    newRow = self.__getUflexPatFoundRow(vlctInst,uflexPatName)
                    self.__digitalInstances.append(newRow)
            elif "IMZE_CKWGLGN2FF_BP" in uflexPatName:
                    newRow = self.__getUflexPatFoundRow(vlctInst,uflexPatName)
                    self.__digitalInstances.append(newRow)
            else: #not found in uflex pats  
                newRow = self.__getUflexPatNotFoundRow(vlctInst)
                self.__digitalInstances.append(newRow)               
        
        
    def __getUflexPatNotFoundRow(self,vlctInst):
                vlctPatName = vlctInst[PATTERN_NAME]
                vlctTestName = vlctInst[TEST_NAME]
                if "USB2_I_CC" in vlctTestName:
                    newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "Pattern run as part of TG_USB2_CCTRIM_ST") 
                elif vlctPatName in self.__patStartedInVba:
                    newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                              "Pattern started in VBT, no datalog line in Ultraflex") 
                elif "PRETRM_" in vlctTestName:
                    newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                              "Trim values vary from device to device  ") 
                elif "ABBOFF" in vlctTestName:
                    newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                              "approved by Adam") 
                elif self.__isScanReadPat(vlctPatName):
                    newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                              "READ/WRITE patterns not separate in UFLEX program  ") 
                else:
                    newRow = ("NO",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctPatName,"","","","")
                return newRow
    
    
    def __getUflexPatFoundRow(self,vlctInst,uflexPatName):    
                vlctPatName = vlctInst[PATTERN_NAME]
                vlctTestName = vlctInst[TEST_NAME]
                uflexTestName = self.__getUflexTestname(vlctTestName)
                uflexInst = self.__getUflexTestInst(uflexTestName,uflexPatName)
                if uflexInst == NOT_FOUND:                      
                    if "USB2_I_CC" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "Pattern run as part of TG_USB2_CCTRIM_ST") 
                    elif vlctPatName in self.__patStartedInVba:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "Pattern started in VBT, no datalog line in Ultraflex") 
                    elif self.__skipInitPats(vlctInst):
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "Init pattern not run on UFlex") 
                    elif "PRETRM_" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "Trim values vary from device to device  ") 
                    elif "SMARTRFLXREAD" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "SMARTRFLXREAD not needed in flow") 
                    elif "HDMI_IRDROP" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "HDMI_IRDROP not needed in flow") 
                    elif "TP_DSIC_HSTX_1_MO_ST" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "Covered as part of UFlex test TG_DSIC_HSTX_EMUL_MO") 
                    elif "POWERSUM_950_ST" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "Test removed from flow") 
                    elif "PTRIM_HDMI_SWC_ST" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "these patterns run under TestGroupName TG_PTRIMHDMI_CHAR0_1_ST") 
                    elif "_1_MO_ST" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "Pattern split in two on UFlex") 
                    elif "PTRIM_VLDO" in vlctTestName:
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "repeated run on VLCT. Run only once on Ultraflex") 
                    elif ("CSIB" in vlctTestName) and ("XBCRXAA" in vlctPatName):
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "VLCT uses CRXAB pat to setup and CRXAA pat to control PM\
                                  stops. UFlex uses CRXAB to do both.") 
                    elif ("CSIC" in vlctTestName) and ("XBCRXAA" in vlctPatName):
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "VLCT uses CRXAC pat to setup and CRXAA pat to control PM\
                                  stops. UFlex uses CRXAC to do both.") 
                    elif self.__isScanReadPat(vlctPatName):
                        newRow = ("YES",vlctInst[LINE_NUM],vlctTestName,vlctPatName,"","","",
                                  "READ/WRITE patterns not separate in UFLEX program") 
                    else:
                        newRow = ("NO",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctPatName,"","","","")
                else:    
                    newRow = ("YES",vlctInst[LINE_NUM],vlctTestName, vlctPatName,
                              uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName,"")
                    try:
                        self.__uflexTestInstances.remove(uflexInst)
                    except:
                        self.__doNothing()
                return newRow
    
    
    def __isScanReadPat(self,vlctPatName):
                hasR = True if "R" in vlctPatName else False
                endIn_B = True if vlctPatName[-3:-1] == "_B" else False
                return hasR and endIn_B
   
    
    def __getUflexTestInst(self,uflexTestName,uflexPatName):
        if uflexPatName == "IATOP_ALL_ACTIVE":
            uflexTestName = "IDDQ_VIDEOPLAYBACK"
        elif uflexPatName == "PMTSMPU":
            uflexTestName = "PTRIM_TS_VNOM"
        elif uflexPatName == "XBUS2A_GPIO_RX":
            uflexTestName = "USB2_GPIORX_ST"
        elif "IATOP_" in uflexPatName: 
            uflexTestName = self.__iatopPairedTests[uflexTestName]
        uflexTestInstances = [inst for inst in self.__uflexTestInstances if uflexTestName in inst]        
        uFlexInst = next((inst for inst in uflexTestInstances if uflexPatName in inst), NOT_FOUND)
        if "BOOT_ATPG_TC_VMAX_ST" in uflexTestName:
            uFlexInst = tuple(["18349","BOOT_ATPG_TC_VMAX_ST","BOOT_ATPG","","","",""])
        return uFlexInst
                

    def __getUflexDigitalTests(self,uflexPatName):
        uflexTestInstances = [inst for inst in self.__uflexTestInstances if uflexPatName in inst]
        for uflexInst in uflexTestInstances:
            if self.__skipInitPats(uflexInst):
                newRow = ("YES","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName,
                          "INIT Pat only run once on VLCT")
                self.__digitalInstances.append(newRow)
                self.__doNothing()
            elif "_SRM" in uflexPatName:
                newRow = ("YES","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName,
                          "subroutine pattern for UFlex")
                self.__digitalInstances.append(newRow)
            elif "MPDSS_CKW0_BZGS_BP" in uflexPatName:
                newRow = ("YES","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName,
                          "write pat only run once on VLCT, before MP_DSS1_RET_TC_OPNO_CORE_ST")
                self.__digitalInstances.append(newRow)
            elif "TP_DSIA_HSTX_1_MO_ST" in uflexPatName:
                newRow = ("YES","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName,
                        "split into two tests on UFlex, TP_DSIA_HSTX_0_MO_ST & TP_DSIA_HSTX_1_MO_ST")
                self.__digitalInstances.append(newRow)
            elif "LPRX_" in uflexPatName:
                newRow = ("YES","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName,
                          "parametric meas in VLCT")
                self.__digitalInstances.append(newRow)
            elif self.__isScanReadPat(uflexPatName):
                newRow = ("YES","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName,
                          "READ/WRITE patterns not separate in UFLEX program") 
            elif "_LPCD" in uflexPatName:
                newRow = ("YES","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName,
                          "parametric meas in VLCT")
                self.__digitalInstances.append(newRow)
            elif not self.__ignoreInst(uflexInst):
                newRow = ("NO","","","",uflexInst[LINE_NUM],uflexInst[TEST_NAME],uflexPatName)
                self.__digitalInstances.append(newRow)
        
        
    def __skipInitPats(self,uflexInst):
        uTest = uflexInst[TEST_NAME]
        uPat = uflexInst[PATTERN_NAME]
        #if ("_V_" in uTest):
            #skipTest = True 
        if (uTest in self.__csiSkipInitTests):
            skipTest = True 
        else:
            skipTest = False 
        skipPat = True if ( uPat in self.__csiInitPats) else False
        return (skipTest and skipPat)
        
    
    def __attachExistingComments(self):
        self.__attachDigitalComments()
        self.__attachAnalogComments()
    
    
    def __attachDigitalComments(self):
        for inst in self.__digitalInstances:
            if inst[VLCT_TEST]:
                vTest = inst[VLCT_TEST]
                vPat = inst[VLCT_PAT]
                commentInstances = [inst for inst in self.__patComments if vTest in inst]
                commentInst = next((inst for inst in commentInstances if vPat in inst), NOT_FOUND)
                if commentInst != NOT_FOUND:
                    self.__addDigCommentToRow(inst,commentInst)                       
            else:
                uTest = inst[UFLEX_TEST]
                uPat = inst[UFLEX_PAT]
                commentInstances = [inst for inst in self.__patComments if uTest in inst]
                commentInst = next((inst for inst in commentInstances if uPat in inst), NOT_FOUND)
                if commentInst != NOT_FOUND:
                    self.__addDigCommentToRow(inst,commentInst)
    
    
    def __attachAnalogComments(self):
        for inst in self.__analogInstances:
            if inst[VLCT_TEST]:
                vTest = inst[VLCT_TEST]
                vPat = inst[VLCT_PAT]
                commentInstances = [inst for inst in self.__analogComments if vTest in inst]
                commentInst = next((inst for inst in commentInstances if vPat in inst), NOT_FOUND)
                if commentInst != NOT_FOUND:
                    self.__addAnalogCommentToRow(inst,commentInst)                       
            else:
                uTest = inst[9]
                uMeas = inst[10]
                commentInstances = [inst for inst in self.__analogComments if uTest in inst]
                commentInst = next((inst for inst in commentInstances if uMeas in inst), NOT_FOUND)
                if commentInst != NOT_FOUND:
                    self.__addAnalogCommentToRow(inst,commentInst)
    
    
    def __addAnalogCommentToRow(self,inst,commentInst):
        NEW_COMMENT = 4
        instList = list(inst)
        instList[15:] = [commentInst[NEW_COMMENT]]
        newRow = tuple(instList)
        self.__analogInstances.remove(inst)
        self.__analogInstances.append(newRow)
   
    
    def __ignoreInst(self,uflexInst):
        if "BUILD_VBBTABLE_FUNC" in uflexInst[TEST_NAME]:
            return True
        else:
            return False
    
    
    def __addDigCommentToRow(self,inst,commentInst):
        NEW_COMMENT = 4
        instList = list(inst)
        instList[7:] = [commentInst[NEW_COMMENT]]
        if "test is parametric" in commentInst[NEW_COMMENT]:
            instList[:1] = ["YES"]
        newRow = tuple(instList)
        self.__digitalInstances.remove(inst)
        self.__digitalInstances.append(newRow)
   
    
    def __getUflexTestname(self,vlctTestname):       
        if vlctTestname in self.__pairedTests: 
            return self.__pairedTests[vlctTestname]
        elif "BLDO2_PRETRM_" in vlctTestname:
            return "BUILD_VBBTABLE_FUNC"
        elif "PTRIM_VLDO1" in vlctTestname:
            return "PTRIMVSLDO1_VNOM"
        elif "PTRIM_VLDO2" in vlctTestname:
            return "PTRIMVSLDO2_VNOM"
        elif "PTRIM_VLDO3" in vlctTestname:
            return "PTRIMVSLDO3_VNOM"
        elif "PTRIM_VLDO4" in vlctTestname:
            return "PTRIMVSLDO4_VNOM"
        elif "TG_PTRIM_PLABE_LOCK" in vlctTestname:
            return "PTRIM_ADPLLABELOCK" + vlctTestname[-4:]
        elif "USB2_I_CC" in vlctTestname:
            return "USB2_I_CC" + vlctTestname[9] +"_MO"
        else:
            return vlctTestname
        
            
    def __getUflexPatname(self,vlctPatname):
        if vlctPatname in self.__pairedPatterns: 
            return self.__pairedPatterns[vlctPatname]
        elif "FBKSTRIM" in vlctPatname:
            return vlctPatname + "_MODIFIED"
        elif vlctPatname[-8:] == "_FSCAN30":
            return vlctPatname[:-8]
        else:
            return vlctPatname
        
            
    def __createAnalogTestsSheet(self):
        self.__analogInstances.sort(key=lambda tup: (tup[VLCT_FLOW] == "",tup[VLCT_FLOW]))
        for testInstance in self.__analogInstances:
            self.__wsAnalog.append(testInstance)
        
    
    def __createContinuityTestsSheet(self):
        self.__analogInstances.sort(key=lambda tup: (tup[VLCT_FLOW] == "",tup[VLCT_FLOW]))
        for testInstance in self.__analogInstances:
            self.__wsAnalog.append(testInstance)
        
    
    def __createDigitalTestsSheet(self):
        digInstances = list(self.__digitalInstances)
        digInstances.sort(key=lambda tup: (tup[UFLEX_FLOW] == "",tup[UFLEX_FLOW]))
        for testInstance in digInstances:
            self.__wsPatterns.append(testInstance)
        
    
    def __getVlctUflexRow(self,vlctInst,uflexInst):
        #check possible ways pattern could be accounted for
        matchingLimits = self.__limitsMatch(vlctInst, uflexInst)
        continuityTest = self.__isContinuityTest(uflexInst)
        firstInstOnly = True if uflexInst[MEAS_NAME] in self.__1stInstOnlyPats else False        
        testMatch = "YES" if any([matchingLimits,continuityTest,firstInstOnly]) else "NO"
        if continuityTest:
            commentString = "UFlex/VLCT use different V/i methodology"
        elif firstInstOnly:
            commentString = "VLCT only datalog 1st inst. Uflex has detailed datalog" 
        else:
            commentString = ""
        newRow = (testMatch,vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctInst[MEAS_NAME],vlctInst[MIN_LIMIT],
                  vlctInst[MAX_LIMIT],vlctInst[MEAS_RSLT],vlctInst[UNITS],uflexInst[LINE_NUM],uflexInst[TEST_NAME],
                  uflexInst[MEAS_NAME],uflexInst[MIN_LIMIT],uflexInst[MAX_LIMIT],uflexInst[MEAS_RSLT],
                  uflexInst[UNITS],commentString)
        return newRow
    

    def __isContinuityTest(self,uflexInst):
        analogContinuity = True if "ANALOGPINSHRTSTEST" in uflexInst[TEST_NAME] else False
        supplyOpens = True if "SUPPLYOPENSTEST" in uflexInst[TEST_NAME] else False
        supplyShorts = True if "SUPPLYSHORTSPRE" in uflexInst[TEST_NAME] else False
        return any([analogContinuity,supplyOpens,supplyShorts])
   
    
    def __getVlctRow(self,vlctInst):
        if "HDMI_IRDROP" in vlctInst[TEST_NAME]:
            newRow = ("YES",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctInst[MEAS_NAME],
                  vlctInst[MIN_LIMIT],vlctInst[MAX_LIMIT],vlctInst[MEAS_RSLT],vlctInst[UNITS],
                  "","","","","","","",
                  "Test not needed for UFlex flow")
        elif vlctInst[MEAS_NAME][:2] == "V_":
            newRow = ("YES",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctInst[MEAS_NAME],
                  vlctInst[MIN_LIMIT],vlctInst[MAX_LIMIT],vlctInst[MEAS_RSLT],vlctInst[UNITS],
                  "","","","","","","",
                  "Functional test which performed retest and datalogged the last failing value")
        elif "ABBOFF" in vlctInst[TEST_NAME]:
            newRow = ("YES",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctInst[MEAS_NAME],
                  vlctInst[MIN_LIMIT],vlctInst[MAX_LIMIT],vlctInst[MEAS_RSLT],vlctInst[UNITS],
                  "","","","","","","",
                  "aproved by Adam")
        elif "PINSHORTSTEST" in vlctInst[TEST_NAME]:
            newRow = ("YES",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctInst[MEAS_NAME],
                  vlctInst[MIN_LIMIT],vlctInst[MAX_LIMIT],vlctInst[MEAS_RSLT],vlctInst[UNITS],
                  "","","","","","","",
                  "Pin Shorts datalogged by group on VLCT and by pin on UFlex")
        elif "PRETRM_" in vlctInst[MEAS_NAME]:
            newRow = ("YES",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctInst[MEAS_NAME],
                  vlctInst[MIN_LIMIT],vlctInst[MAX_LIMIT],vlctInst[MEAS_RSLT],vlctInst[UNITS],
                  "","","","","","","",
                  "exact trim values change from device to device")
        elif "ANALOGPINOPENSTEST" in vlctInst[MEAS_NAME]:
            newRow = ("YES",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctInst[MEAS_NAME],
                  vlctInst[MIN_LIMIT],vlctInst[MAX_LIMIT],vlctInst[MEAS_RSLT],vlctInst[UNITS],
                  "","","","","","","",
                  "opens test not run on UFlex")
        else:
            newRow = ("NO",vlctInst[LINE_NUM],vlctInst[TEST_NAME],vlctInst[MEAS_NAME],vlctInst[MIN_LIMIT],
                  vlctInst[MEAS_RSLT],vlctInst[MAX_LIMIT],vlctInst[UNITS],"","","","","","","","")                    
        return newRow
    

    def __limitsMatch(self,vlctInst,uflexInst):
        vlctInst = self.__convertToFloatInst(vlctInst)
        uflexInst = self.__convertToFloatInst(uflexInst)
        if (vlctInst[MIN_LIMIT] == uflexInst[MIN_LIMIT]) and (vlctInst[MAX_LIMIT] == uflexInst[MAX_LIMIT]):
            return True
        else:
            return False
    
    def __convertToFloatInst(self,inst):
        try:
            minLim = float(inst[MIN_LIMIT])
        except:
            minLim = None
        try:
            maxLim = float(inst[MAX_LIMIT])
        except:
            maxLim = None
        newInst = (inst[LINE_NUM],inst[TEST_NAME],inst[PATTERN_NAME],inst[MEAS_NAME],
                   minLim,inst[MEAS_RSLT],maxLim,inst[UNITS])
        return newInst
    
    
    def __createVlctDatalog(self):
        self.__vlctTestInstances.sort(key=lambda tup: tup[FLOW_NUM])
        for testInstance in self.__vlctTestInstances:
            if "PTRIM_ADPLLABELOCK1_ST" in testInstance[TEST_NAME]:
                self.__doNothing()
            self.__wsVlctDatalog.append(testInstance)
        
    
    def __createUflexDatalog(self):
        self.__uflexTestInstances.sort(key=lambda tup: tup[FLOW_NUM])
        for testInstance in self.__uflexTestInstances:
            self.__wsUflexDatalog.append(testInstance)
        
    
    def __createExcelTable(self,ws,tableName):
        red_font = Font(color='00FF0000', italic=True)
        #create a table with pattern data
        refCells = "A1:" + chr(64 + ws.max_column) + str(ws.max_row)
        newTable = Table(displayName=tableName, ref=refCells)                      
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium13", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        newTable.tableStyleInfo = style
        ws.add_table(newTable)
        freezeRowsAboveCell = ws['B2']
        ws.freeze_panes = freezeRowsAboveCell
        for row in ws.rows:
            if row[0]._value  == "NO":
                for cell in row:
                    cell.font = red_font
            
        
    
    def __doNothing(self):
        import time    
        time.sleep(0)
        
#used to test class using runs as compareVlctUflexDatalogsClass
if __name__ == '__main__':
    from time import sleep
    
    compareDatalogs = compareVlctUflexDatalogsClass()    
    compareDatalogs.parseVlctDatalog("vlctDatalog.txt")
    compareDatalogs.parseUflexDatalog("FT1_FullDatalog_20180205.txt")
    compareDatalogs.comparePatterns()
    compareDatalogs.compareAnalogTests()
    compareDatalogs.saveWorkbook()
    
    sleep(0)  
        